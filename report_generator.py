"""
Excel report generator for firewall log analysis results.
"""

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os


def generate_excel_report(df: pd.DataFrame, output_path: str):
    """Export firewall events to a formatted Excel file."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: All events
        df.to_excel(writer, sheet_name='All Events', index=False)

        # Sheet 2: Denied only
        denied = df[df['action'].isin(['DENY', 'DROP', 'BLOCK'])]
        denied.to_excel(writer, sheet_name='Blocked Events', index=False)

        # Sheet 3: Top IPs summary
        top_ips = denied['src_ip'].value_counts().reset_index()
        top_ips.columns = ['Source IP', 'Block Count']
        top_ips.to_excel(writer, sheet_name='Top Blocked IPs', index=False)

        # Auto-size columns on all sheets
        for sheet in writer.sheets.values():
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    print(f"Excel report generated: {output_path}")
